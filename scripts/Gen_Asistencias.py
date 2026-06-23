import pandas as pd
import unicodedata
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.worksheet.page import PageMargins
from tkinter import filedialog, messagebox
import tkinter as tk
import os
import locale
import os
from tkinter import messagebox
from datetime import datetime
import subprocess
import tkinter as tk
import webbrowser

# Configuración regional para nombres de días en español
locale.setlocale(locale.LC_TIME, 'Spanish_Mexico.1252')
root = tk.Tk()
root.withdraw()
def mostrar_ventana_licencia():
    ventana = tk.Toplevel()
    ventana.title("Número de usos excedido")
    ventana.geometry("450x250")
    ventana.configure(bg="#f0f0f0")
    ventana.resizable(False, False)
    # ===== TÍTULO EN NEGRITA Y GRANDE =====
    titulo = tk.Label(
        ventana,
        text="Ya no puedes generar más archivos.",
        font=("Helvetica", 16, "bold"),
        fg="#800000",
        bg="#f0f0f0"
    )
    titulo.pack(pady=(20, 10))
    # ===== MENSAJE PRINCIPAL =====
    mensaje = (
        "Ya has utilizado todas tus pruebas gratuitas.\n\n"
        "Para seguir usando el programa, contacta a tu proveedor de servicios\n"
        "y solicita una licencia completa."
    )
    etiqueta = tk.Label(
        ventana,
        text=mensaje,
        font=("Segoe UI", 10),
        bg="#f0f0f0",
        justify="left",
        wraplength=460
    )
    etiqueta.pack(padx=20, pady=5)
    # ===== BOTONES ENLACES =====
    def abrir_correo():
        webbrowser.open("mailto:victormontano.j@gmail.com")
    def abrir_web():
        webbrowser.open("https://clerdevs-code.onrender.com")
    btn_correo = tk.Button(
      ventana,
      text="✉ Contactar por correo",
      command=abrir_correo,
      fg="white",
      bg="#0078D7",
      activebackground="#005a9e",
      relief="raised",
      font=("Segoe UI", 10),
      cursor="hand2"
  )
    btn_correo.pack(pady=8)

    btn_web = tk.Button(
      ventana,
      text="🌐 Visitar sitio web",
      command=abrir_web,
      fg="white",
      bg="#28a745",
      activebackground="#1e7e34",
      relief="raised",
      font=("Segoe UI", 10),
      cursor="hand2"
  )
    btn_web.pack(pady=4)

  # ===== BOTÓN CERRAR =====
    cerrar = tk.Button(
      ventana,
      text="Cerrar",
      command=ventana.destroy,
      font=("Segoe UI", 9),
      bg="#cccccc",
      relief="flat"
  )
    cerrar.pack(pady=20)

    ventana.transient()
    ventana.grab_set()
    ventana.mainloop()
# ===== ARCHIVO DE ENTRADA =====
ARCHIVO_ENTRADAS = filedialog.askopenfilename(
  title="Selecciona el archivo de registro",
  filetypes=[("Archivos Excel", "*.xlsx *.xls")]
)
if not ARCHIVO_ENTRADAS:
  messagebox.showerror("Error", "No se seleccionó ningún archivo.")
  exit()
NOMBRE_SALIDA = filedialog.asksaveasfilename(
  defaultextension=".xlsx",
  filetypes=[("Archivos Excel", "*.xlsx")],
  title="Guardar como",
  initialfile=f"ASISTENCIAS_JASANA_{datetime.today().strftime('%d-%m-%Y')}.xlsx"
)

if not NOMBRE_SALIDA:
  messagebox.showinfo("Cancelado", "El guardado fue cancelado por el usuario.")
  exit()
# ===== CARGA DE DATOS =====
df = pd.read_excel(ARCHIVO_ENTRADAS)
for col in ['Empleado', 'Entrada', 'Salida', 'Horas trabajadas', 'Horas extra', 'Departamento']:
  if col not in df.columns:
    raise ValueError(f"Falta la columna '{col}' en el archivo Excel.")
df['Entrada'] = pd.to_datetime(df['Entrada'], errors='coerce')
df['Salida'] = pd.to_datetime(df['Salida'], errors='coerce')
empleados = df['Empleado'].dropna().unique()
# ===== CREACIÓN DE ARCHIVO NUEVO =====
wb = Workbook()
del wb['Sheet']
# Estilos
color_fondo_encabezado = PatternFill(start_color="CCC0DA", end_color="CCC0DA", fill_type="solid")
color_fuente_encabezado = Font(bold=True)
border = Border(
  left=Side(style="thin", color="000000"),
  right=Side(style="thin", color="000000"),
  top=Side(style="thin", color="000000"),
  bottom=Side(style="thin", color="000000")
)
for empleado in empleados:
  df_emp = df[df['Empleado'] == empleado].dropna(subset=['Entrada', 'Salida'])
  if df_emp.empty:
    continue
  registro_dias = defaultdict(list)
  for _, row in df_emp.iterrows():
    fecha = row['Entrada'].date()
    registro_dias[fecha].append({'Entrada': row['Entrada'], 'Salida': row['Salida']})
  departamento = unicodedata.normalize('NFKD', df_emp['Departamento'].iloc[0].lower()).encode('ASCII', 'ignore').decode('utf-8')
  # Obtener fechas mínimas y máximas de entrada
  fecha_min = df_emp['Entrada'].min().date()
  fecha_max = df_emp['Entrada'].max().date()

  # Ajustar fecha de inicio al viernes anterior (o igual si es viernes)
  fecha_inicio = fecha_min - timedelta(days=(fecha_min.weekday() - 4) % 7)

  # Ajustar fecha de fin al jueves siguiente (o igual si es jueves)
  fecha_fin = fecha_max + timedelta(days=(3 - fecha_max.weekday()) % 7)

  # Crear rango completo de días de viernes a jueves
  dias_laborales = pd.date_range(start=fecha_inicio, end=fecha_fin, freq='D')

  nombre_hoja = empleado[:31]
  ws = wb.create_sheet(title=nombre_hoja)
  ws.insert_rows(1, amount=5)
  logo = Image("Logo.png")
  logo.width = 120
  logo.height = 45
  logo.anchor = "H1"
  ws.add_image(logo)

  ws.merge_cells('A1:H1')
  cell = ws['A1']
  cell.value = "REGISTRO DE ENTRADAS Y SALIDAS"
  cell.font = Font(bold=True, size=14)
  cell.alignment = Alignment(horizontal='center', vertical='center')

  ws['A2'].value = "NO."
  ws['A2'].font = Font(bold=True)
  ws['A2'].fill = color_fondo_encabezado
  ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

  ws['B2'].value = empleados.tolist().index(empleado) + 1
  ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

  ws['C2'].value = "NOMBRE"
  ws['C2'].font = Font(bold=True)
  ws['C2'].fill = color_fondo_encabezado
  ws['C2'].alignment = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('D2:G2')
  ws['D2'].value = empleado
  ws['D2'].alignment = Alignment(horizontal='left', vertical='center')

  ws.merge_cells('A3:B3')
  ws['A3'].value = "DEL"
  ws['A3'].font = Font(bold=True)
  ws['A3'].fill = color_fondo_encabezado
  ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('C3:D3')
  ws['C3'].value = fecha_inicio.strftime("%d/%m/%Y")
  ws['C3'].alignment = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('E3:F3')
  ws['E3'].value = "AL"
  ws['E3'].font = Font(bold=True)
  ws['E3'].fill = color_fondo_encabezado
  ws['E3'].alignment = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('G3:H3')
  ws['G3'].value = fecha_fin.strftime("%d/%m/%Y")
  ws['G3'].alignment = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('A4:D4')
  ws['A4'].value = "HORARIO"
  ws['A4'].font = Font(bold=True)
  ws['A4'].fill = color_fondo_encabezado
  ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

  ws.merge_cells('E4:H4')
  if departamento in ("administration", "administración"):
    ws['E4'].value = "DE 8:00 A 18:00 HRS"
    hora_salida_normal = datetime.strptime("18:00", "%H:%M").time()
    hora_entrada_limite = datetime.strptime("08:06", "%H:%M").time()
  elif departamento in ("logística", "logistica"):
    ws['E4'].value = "DE 8:30 A 18:00 HRS"
    hora_salida_normal = datetime.strptime("18:00", "%H:%M").time()
    hora_entrada_limite = datetime.strptime("08:36", "%H:%M").time()
  elif departamento in ("sistemas", "sistemas"):
    ws['E4'].value = "DE 13:00 A 18:00 HRS"
    hora_salida_normal = datetime.strptime("18:00", "%H:%M").time()
    hora_entrada_limite = datetime.strptime("13:06", "%H:%M").time()
  elif departamento in ("pasante", "pasante"):
    ws['E4'].value = "DE 12:00 A 18:00 HRS"
    hora_salida_normal = datetime.strptime("18:00", "%H:%M").time()
    hora_entrada_limite = datetime.strptime("12:06", "%H:%M").time()
  else:
    ws['E4'].value = "DE 8:00 A 17:30 HRS"
    hora_salida_normal = datetime.strptime("17:30", "%H:%M").time()
    hora_entrada_limite = datetime.strptime("08:06", "%H:%M").time()


  ws['E4'].alignment = Alignment(horizontal='center', vertical='center')

  ws.append(["DIA", "ENTRADA", "DESAYUNO","", "COMIDA", "", "SALIDA", "OBSERVACIONES"])
  ws.append(["", "", "SALIDA", "ENTRADA", "SALIDA", "ENTRADA", "", ""])
  ws.merge_cells('C7:D7')
  ws.merge_cells('E7:F7')
  for row in ws.iter_rows(min_row=7, max_row=8, min_col=1, max_col=8):
    for cell in row:
      cell.fill = color_fondo_encabezado
      cell.font = color_fuente_encabezado
      cell.border = border
      cell.alignment = Alignment(horizontal='center', vertical='center')
  retardos = 0
  faltas = 0
  minutos_extra = 0

  for idx, dia in enumerate(dias_laborales, start=9):
    fecha = dia.date()
    entrada = salida_desayuno = entrada_desayuno = salida_comida = entrada_comida = salida = None
    observacion = ""
    if fecha in registro_dias:
      eventos = []
      for registro in registro_dias[fecha]:
        eventos.append(registro['Entrada'])
        eventos.append(registro['Salida'])
      eventos = sorted([e for e in eventos if pd.notna(e)])
      for e in eventos:
        t = e.time()
        if not entrada and t <= datetime.strptime("09:00", "%H:%M").time():
          entrada = e
        elif not salida_desayuno and t <= datetime.strptime("11:30", "%H:%M").time():
          salida_desayuno = e
        elif not entrada_desayuno and t <= datetime.strptime("13:00", "%H:%M").time():
          entrada_desayuno = e
        elif not salida_comida and t <= datetime.strptime("15:30", "%H:%M").time():
            salida_comida = e
        elif not entrada_comida and t <= datetime.strptime("17:00", "%H:%M").time():
          entrada_comida = e
        else:
          salida = e
      if entrada and entrada.time() > hora_entrada_limite:
        observacion = "RETARDO"
        retardos += 1
      if salida and salida.time() > hora_salida_normal:
        delta = datetime.combine(datetime.min, salida.time()) - datetime.combine(datetime.min, hora_salida_normal)
        minutos_extra += int(delta.total_seconds() / 60)
      else:
      # Si no hay registro para ese día
        if fecha.weekday() in [0, 1, 2, 3, 4]:  # Lunes (0) a Viernes (4)
          observacion = "FALTA"
          faltas += 1
        else:
          observacion = ""  # Sábado y domingo, sin registro, no se cuentan como faltas

    ws.append([
      fecha.strftime('%A, %d').capitalize(),
      entrada.strftime("%H:%M") if entrada else "",
      salida_desayuno.strftime("%H:%M") if salida_desayuno else "",
      entrada_desayuno.strftime("%H:%M") if entrada_desayuno else "",
      salida_comida.strftime("%H:%M") if salida_comida else "",
      entrada_comida.strftime("%H:%M") if entrada_comida else "",
      salida.strftime("%H:%M") if salida else "",
      observacion
    ])
    for col in range(1, 9):
      cell = ws.cell(row=idx, column=col)
      cell.border = border
      cell.alignment = Alignment(horizontal='center', vertical='center')
  fila_resumen = idx + 2
  ws.append([])
  # Conversión de minutos a HH:MM
  if minutos_extra > 60:
    horas = minutos_extra // 60
    minutos = minutos_extra % 60
    tiempo_extra = f"{horas}:{minutos:02d}"
  else:
    tiempo_extra = "0:00"
  ws.append([f"Retardos", retardos, "", "Faltas", faltas,"", "Hrs.Extra:", tiempo_extra])
  for col in range(1, 9):
    cell = ws.cell(row=fila_resumen, column=col)
    cell.font = Font(bold=True)
    cell.border = border
    cell.alignment = Alignment(horizontal='center', vertical='center')
  fila_firma = fila_resumen + 3
  borde_firma = Side(style="thick", color="000000")
  for col in range(2, 8):
    cell = ws.cell(row=fila_firma, column=col)
    cell.border = Border(bottom=borde_firma)
    cell.alignment = Alignment(horizontal='center')
    cell.value = ""
  fila_nombre = fila_firma + 1
  ws.merge_cells(start_row=fila_nombre, start_column=3, end_row=fila_nombre, end_column=6)
  cell_nombre = ws.cell(row=fila_nombre, column=3)
  cell_nombre.value = empleado
  cell_nombre.alignment = Alignment(horizontal='center')

  fila_departamento = fila_nombre + 1
  ws.merge_cells(start_row=fila_departamento, start_column=3, end_row=fila_departamento, end_column=6)
  cell_departamento = ws.cell(row=fila_departamento, column=3)
  cell_departamento.value = df_emp['Departamento'].iloc[0]
  cell_departamento.alignment = Alignment(horizontal='center')
  cell_departamento.font = Font(italic=True, size=10)

  for col in [1, 2, 8]:
    cell = ws.cell(row=fila_nombre, column=col)
    cell.value = ""
    cell.alignment = Alignment(horizontal='center')
  anchos = [13, 9.5, 9.5, 9.5, 9.5, 9.5, 9.5, 21]
  for i, ancho in enumerate(anchos, start=1):
    ws.column_dimensions[get_column_letter(i)].width = ancho
  ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
# ===== GUARDAR ARCHIVO =====
wb.save(NOMBRE_SALIDA)
def mostrar_ventana_completado(ruta_archivo):
    ventana = tk.Toplevel()
    ventana.title("¡Tarea Completada!")
    ventana.geometry("500x180")
    ventana.configure(bg="#f0f0f0")
    ventana.resizable(False, False)

    # Etiqueta de mensaje
    mensaje = tk.Label(
        ventana,
        text=f"Archivo Excel guardado correctamente en:\n{ruta_archivo}",
        font=("Segoe UI", 10),
        bg="#f0f0f0",
        wraplength=480,
        justify="left"
    )
    mensaje.pack(pady=(20, 10), padx=20)

    # Función para abrir el archivo
    def abrir_archivo():
        try:
            os.startfile(ruta_archivo)  # Solo funciona en Windows
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{str(e)}")

    # Botón para abrir archivo
    btn_abrir = tk.Button(
        ventana,
        text="📂 Abrir archivo",
        command=abrir_archivo,
        font=("Segoe UI", 10),
        bg="#0078D7",
        fg="white",
        activebackground="#005a9e",
        relief="raised",
        cursor="hand2"
    )
    btn_abrir.pack(pady=8)

    # Botón para cerrar ventana
    btn_cerrar = tk.Button(
        ventana,
        text="Cerrar",
        command=ventana.destroy,
        font=("Segoe UI", 9),
        bg="#cccccc",
        relief="flat"
    )
    btn_cerrar.pack(pady=(0, 15))

    ventana.transient()
    ventana.grab_set()
    ventana.mainloop()

# Llamar a la función después de guardar
mostrar_ventana_completado(NOMBRE_SALIDA)
